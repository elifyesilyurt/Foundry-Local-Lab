#!/usr/bin/env python3
"""
Microsoft EcoRAG — Benchmark Test Set 2 Runner
Runs Microsoft_EcoRAG_Benchmark_Test_Set_2.xlsx against the live pipeline
and produces a detailed results report.
"""

import os
import sys
import re
import json
import time
import sqlite3
import unicodedata
from collections import defaultdict
from datetime import datetime

import numpy as np
import openpyxl

# ── Paths ─────────────────────────────────────────────────────────────────────
REPO_DIR   = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH  = os.path.join(REPO_DIR, "Microsoft_EcoRAG_Benchmark_Test_Set_2.xlsx")
DB_PATH    = os.path.join(REPO_DIR, "rag_storage.db")
OUT_JSON   = os.path.join(REPO_DIR, "benchmark_set2_results.json")
OUT_REPORT = os.path.join(REPO_DIR, "BENCHMARK_SET2_REPORT.md")

sys.path.insert(0, REPO_DIR)

# ── Config (mirrors app.py) ───────────────────────────────────────────────────
EMBEDDING_MODEL_NAME = "nomic-ai/nomic-embed-text-v1.5"
FOUNDRY_BASE_URL     = "http://127.0.0.1:62095"
MODEL_NAME           = "Phi-4-mini-instruct-generic-gpu"
RELATIVE_DROP_RATIO  = 0.70
MAX_K                = 6
MIN_SCORE_FLOOR      = 0.15

# ── Lazy globals ──────────────────────────────────────────────────────────────
_embed_model = None
_conn        = None

def get_embed_model():
    global _embed_model
    if _embed_model is None:
        print("  [init] Loading embedding model…", flush=True)
        from sentence_transformers import SentenceTransformer
        _embed_model = SentenceTransformer(EMBEDDING_MODEL_NAME, trust_remote_code=True)
    return _embed_model

def get_conn():
    global _conn
    if _conn is None:
        _conn = sqlite3.connect(DB_PATH)
        _conn.row_factory = sqlite3.Row
    return _conn

# ── Helpers ───────────────────────────────────────────────────────────────────
def normalize_str(s: str) -> str:
    return unicodedata.normalize("NFD", s.lower().strip())

def embed_query(text: str) -> np.ndarray:
    model = get_embed_model()
    vec = model.encode(f"search_query: {text}", normalize_embeddings=True)
    return vec.astype(np.float32)

_docs_cache = None   # list of dicts with pre-parsed numpy embeddings

def _load_docs_cache():
    global _docs_cache
    if _docs_cache is not None:
        return _docs_cache
    conn = get_conn()
    rows = conn.execute("SELECT id, year, title, content, embedding FROM documents").fetchall()
    _docs_cache = []
    for r in rows:
        blob = r["embedding"]
        if isinstance(blob, (bytes, bytearray)):
            vec = np.frombuffer(blob, dtype=np.float32).copy()
        else:
            vec = np.array(json.loads(blob), dtype=np.float32)
        _docs_cache.append({
            "id": r["id"], "year": r["year"], "title": r["title"],
            "content": r["content"], "vec": vec
        })
    print(f"  [init] Loaded {len(_docs_cache)} document embeddings into memory.", flush=True)
    return _docs_cache

def search_context_hybrid(query: str):
    docs = _load_docs_cache()
    if not docs:
        return [], 0.0

    q_vec  = embed_query(query)
    q_norm = normalize_str(query)
    scores = []
    for doc in docs:
        sim         = float(np.dot(q_vec, doc["vec"]))
        words       = q_norm.split()
        c_text      = normalize_str(doc["content"])
        match_count = sum(1 for w in words if len(w) > 3 and w in c_text)
        hybrid_score = sim + (0.10 * match_count)
        scores.append({
            "id": doc["id"], "year": doc["year"], "title": doc["title"],
            "content": doc["content"], "score": hybrid_score
        })

    scores.sort(key=lambda x: x["score"], reverse=True)
    if not scores:
        return [], 0.0

    max_score = scores[0]["score"]
    if max_score < MIN_SCORE_FLOOR:
        return [], max_score

    # Year-Stratified for multi-year queries
    found_years = re.findall(r'\b(2024|2025|2026)\b', query)
    found_fys = re.findall(r'\bfy\s*(2[0-6])\b', query.lower())

    is_multi_year = bool(
        len(found_years) >= 2 or
        len(found_fys) >= 2 or
        (len(found_years) >= 1 and len(found_fys) >= 1) or
        any(w in q_norm for w in [
            "uc yillik", "3 yillik", "tarihsel", "karsilastir",
            "gelisim", "trajectory", "multi-year", "across the",
            "across reports", "trend", "fark", "ilerle",
            "cross-document", "all three", "all reports"
        ])
    )

    if is_multi_year:
        y2024 = [s for s in scores if "2024" in str(s.get("year","")) or "2024" in str(s.get("title",""))][:2]
        y2025 = [s for s in scores if "2025" in str(s.get("year","")) or "2025" in str(s.get("title",""))][:2]
        y2026 = [s for s in scores if "2026" in str(s.get("year","")) or "2026" in str(s.get("title",""))][:2]
        
        if ('23' in found_fys or '2024' in found_years) and ('25' in found_fys or '2026' in found_years) and '24' not in found_fys and '2025' not in found_years:
            y2024_top3 = [s for s in scores if "2024" in str(s.get("year","")) or "2024" in str(s.get("title",""))][:3]
            y2026_top3 = [s for s in scores if "2026" in str(s.get("year","")) or "2026" in str(s.get("title",""))][:3]
            stratified = y2026_top3 + y2024_top3
        else:
            stratified = y2026 + y2025 + y2024

        if len(stratified) >= 3:
            return stratified, max_score

    cutoff   = max_score * RELATIVE_DROP_RATIO
    filtered = [item for item in scores[:MAX_K] if item["score"] >= cutoff]
    return filtered, max_score

def query_foundry(system_prompt: str, user_prompt: str) -> str:
    import requests
    url = f"{FOUNDRY_BASE_URL}/v1/chat/completions"
    payload = {
        "model": MODEL_NAME,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user",   "content": user_prompt}
        ],
        "temperature": 0.0,
        "max_tokens": 512,
        "stream": False,
    }
    try:
        r = requests.post(url, json=payload, timeout=60)
        r.raise_for_status()
        return r.json()["choices"][0]["message"]["content"].strip()
    except Exception as e:
        return f"[ERROR: {e}]"

SYNTHESIS_PROMPT = """Sen kıdemli bir Sürdürülebilirlik Analistisin.
Aşağıda verilen rapor bölümlerini kullanarak soruyu açık, doğru ve öz biçimde yanıtla.
Bilgi yoksa "Bu bilgi raporlarda yer almamaktadır." de — asla tahmin etme."""

def build_user_prompt(question: str, chunks) -> str:
    ctx = "\n\n---\n\n".join(
        f"[{c['title']}]\n{c['content']}" for c in chunks[:4]
    )
    return f"Bağlam:\n{ctx}\n\nSoru: {question}"

# ── Scoring ───────────────────────────────────────────────────────────────────
ADVERSARIAL_REJECTION_PHRASES = [
    "yer almamaktadır", "bulunamadı", "bilgi yok", "mevcut değil",
    "raporlarda yer alm", "bu bilgi", "doğrulayamam", "bulamıyorum",
    "belirtilmemiş", "açıklanmamış", "not found", "not available",
    "not mentioned", "not provided", "cannot confirm", "no information",
    "this information", "not in the report"
]

def compute_faithfulness(answer: str, chunks) -> float:
    """Simple token overlap faithfulness score."""
    if not chunks or not answer:
        return 0.0
    ctx_text = " ".join(c["content"] for c in chunks)
    answer_words = set(normalize_str(answer).split())
    ctx_words    = set(normalize_str(ctx_text).split())
    answer_words = {w for w in answer_words if len(w) > 3}
    if not answer_words:
        return 0.5
    overlap = len(answer_words & ctx_words) / len(answer_words)
    return round(min(overlap, 1.0), 4)

def score_answer(question: str, expected: str, answer: str,
                 category: str, chunks) -> dict:
    """
    Returns dict with: passed (bool), faithfulness (float), notes (str)
    """
    exp_norm = normalize_str(expected)
    ans_norm = normalize_str(answer)

    # ── Adversarial category: model must REJECT ──────────────────────────────
    if "adversarial" in category.lower() or "negative" in category.lower():
        rejected = any(p in ans_norm for p in ADVERSARIAL_REJECTION_PHRASES)
        return {
            "passed": rejected,
            "faithfulness": 1.0 if rejected else 0.0,
            "notes": "Correct rejection" if rejected else "HALLUCINATION — should have rejected"
        }

    # ── Numerical / PAL: check if key numbers appear in answer ───────────────
    if "numerical" in category.lower() or "pal" in category.lower() or "trend" in category.lower():
        # Extract numbers from expected
        exp_numbers = re.findall(r'[\d,\.]+', expected)
        if exp_numbers:
            hits = sum(1 for n in exp_numbers if n.replace(",","").replace(".","") in ans_norm.replace(",","").replace(".",""))
            ratio = hits / len(exp_numbers)
            passed = ratio >= 0.5
        else:
            # No numbers, check keyword overlap
            exp_kw = [w for w in exp_norm.split() if len(w) > 4][:5]
            hits = sum(1 for w in exp_kw if w in ans_norm)
            passed = hits >= max(1, len(exp_kw) // 2)
        faith = compute_faithfulness(answer, chunks)
        return {"passed": passed, "faithfulness": faith, "notes": f"{hits}/{len(exp_numbers if exp_numbers else exp_kw)} key values matched"}

    # ── General: keyword overlap scoring ─────────────────────────────────────
    # Extract meaningful keywords from expected answer
    stop_words = {"ve", "ile", "bir", "bu", "de", "da", "the", "and", "of",
                  "in", "is", "are", "was", "to", "for", "that", "it", "as",
                  "has", "have", "been", "by", "an", "with", "from", "at", "on"}
    exp_keywords = [w for w in exp_norm.split() if len(w) > 4 and w not in stop_words][:12]

    if not exp_keywords:
        exp_keywords = [w for w in exp_norm.split() if len(w) > 2][:8]

    if not exp_keywords:
        faith = compute_faithfulness(answer, chunks)
        return {"passed": True, "faithfulness": faith, "notes": "No keywords to check"}

    hits = sum(1 for w in exp_keywords if w in ans_norm)
    ratio = hits / len(exp_keywords)
    passed = ratio >= 0.40   # 40% keyword match threshold
    faith  = compute_faithfulness(answer, chunks)
    return {
        "passed": passed,
        "faithfulness": faith,
        "notes": f"{hits}/{len(exp_keywords)} keywords matched ({ratio:.0%})"
    }

# ── Main runner ───────────────────────────────────────────────────────────────
def run_benchmark():
    print(f"\n{'='*72}")
    print("  Microsoft EcoRAG — Benchmark Test Set 2 Runner")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*72}\n")

    # Load questions
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    questions = []
    for row in range(2, ws.max_row + 1):
        qid      = ws.cell(row, 1).value
        category = ws.cell(row, 2).value
        question = ws.cell(row, 3).value
        expected = ws.cell(row, 4).value
        source   = ws.cell(row, 5).value
        if question:
            questions.append({
                "qid": str(qid), "category": str(category),
                "question": str(question), "expected": str(expected or ""),
                "source": str(source or "")
            })

    total = len(questions)
    print(f"  Loaded {total} questions from XLSX\n")

    # Pre-load embedding model and docs cache
    get_embed_model()
    _load_docs_cache()
    print()

    results   = []
    cat_stats = defaultdict(lambda: {"total": 0, "passed": 0, "faith_sum": 0.0, "lat_sum": 0.0})

    for i, q in enumerate(questions, 1):
        t0 = time.time()
        qid      = q["qid"]
        category = q["category"]
        question = q["question"]
        expected = q["expected"]

        # RAG search
        chunks, max_score = search_context_hybrid(question)

        # Generate answer
        if chunks:
            user_prompt = build_user_prompt(question, chunks)
            answer = query_foundry(SYNTHESIS_PROMPT, user_prompt)
        else:
            answer = "Bu bilgi raporlarda yer almamaktadır."

        latency = time.time() - t0

        # Score
        scoring = score_answer(question, expected, answer, category, chunks)
        passed  = scoring["passed"]
        faith   = scoring["faithfulness"]
        notes   = scoring["notes"]

        # Category key
        cat_key = category.strip()
        cat_stats[cat_key]["total"]    += 1
        cat_stats[cat_key]["passed"]   += int(passed)
        cat_stats[cat_key]["faith_sum"] += faith
        cat_stats[cat_key]["lat_sum"]  += latency

        result = {
            "qid": qid, "category": cat_key, "question": question,
            "expected": expected[:200], "answer": answer[:400],
            "passed": passed, "faithfulness": faith,
            "latency_s": round(latency, 4), "max_score": round(max_score, 4),
            "chunks_found": len(chunks), "notes": notes
        }
        results.append(result)

        status = "✅" if passed else "❌"
        print(f"  [{i:>3}/{total}] {status} Q{qid:>3} [{cat_key[:28]:<28}] {latency:.2f}s  {notes[:40]}")

    # ── Summary ───────────────────────────────────────────────────────────────
    total_passed = sum(1 for r in results if r["passed"])
    total_time   = sum(r["latency_s"] for r in results)
    avg_latency  = total_time / total if total else 0
    accuracy     = total_passed / total * 100 if total else 0
    avg_faith    = sum(r["faithfulness"] for r in results) / total if total else 0

    summary = {
        "total_questions": total,
        "total_passed": total_passed,
        "overall_accuracy": round(accuracy, 2),
        "total_time_seconds": round(total_time, 2),
        "avg_latency_seconds": round(avg_latency, 4),
        "avg_faithfulness": round(avg_faith, 4),
        "run_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "category_breakdown": {
            k: {
                "total": v["total"],
                "passed": v["passed"],
                "accuracy": round(v["passed"] / v["total"] * 100, 2) if v["total"] else 0,
                "avg_faithfulness": round(v["faith_sum"] / v["total"], 4) if v["total"] else 0,
                "avg_latency": round(v["lat_sum"] / v["total"], 4) if v["total"] else 0,
            }
            for k, v in cat_stats.items()
        }
    }

    # Save JSON
    output = {"summary": summary, "category_breakdown": summary["category_breakdown"], "questions": results}
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"\n  Results saved → {OUT_JSON}")

    # ── Markdown Report ───────────────────────────────────────────────────────
    cat_order = [
        "Factual / Retrieval Accuracy",
        "Numerical, Trend & PAL",
        "Cross-Document Reasoning",
        "Adversarial / Negative Rejection",
        "Language, Format & Edge-Case",
    ]

    with open(OUT_REPORT, "w", encoding="utf-8") as f:
        f.write(f"# 📊 Microsoft EcoRAG — Benchmark Test Set 2 Raporu\n\n")
        f.write(f"**Tarih:** {summary['run_date']}  \n")
        f.write(f"**Kaynak:** `Microsoft_EcoRAG_Benchmark_Test_Set_2.xlsx` (500 Soru)  \n")
        f.write(f"**Model:** `{MODEL_NAME}` @ Foundry Local | **Embedding:** `nomic-embed-text-v1.5`  \n")
        f.write(f"**Mimari:** Year-Stratified RAG + PAL Deterministik Motor | **Sıcaklık:** `0.0`\n\n")
        f.write("---\n\n")
        f.write("## 🌟 1. Yönetici Özeti\n\n")
        f.write("| Metrik | Değer | Hedef | Durum |\n")
        f.write("|---|---|---|---|\n")
        f.write(f"| **Toplam Soru** | **{total}** | 500 | ✅ |\n")
        f.write(f"| **Genel Doğruluk** | **%{accuracy:.2f} ({total_passed}/{total})** | >%85 | {'🏆 Üstün' if accuracy >= 91 else ('✅ Başarılı' if accuracy >= 85 else '⚠️ Geliştirilmeli')} |\n")
        f.write(f"| **Ort. Gecikme** | **{avg_latency*1000:.0f} ms / soru** | <1000 ms | ⚡ |\n")
        f.write(f"| **Ort. Sadakat (Faithfulness)** | **{avg_faith:.4f}** | >0.60 | {'✅' if avg_faith >= 0.60 else '⚠️'} |\n")
        f.write(f"| **Toplam Süre** | **{total_time:.1f} sn** | — | — |\n\n")
        f.write("---\n\n")
        f.write("## 📈 2. Kategori Bazlı Sonuçlar\n\n")
        f.write("| Kategori | Soru | Başarılı | Doğruluk | Sadakat | Ort. Gecikme |\n")
        f.write("|---|---|---|---|---|---|\n")

        for cat in cat_order:
            if cat in summary["category_breakdown"]:
                s = summary["category_breakdown"][cat]
                emoji = "🏆" if s["accuracy"] == 100 else ("✅" if s["accuracy"] >= 85 else "⚠️")
                f.write(f"| **{cat}** | {s['total']} | {s['passed']} | {emoji} **%{s['accuracy']:.2f}** | {s['avg_faithfulness']:.4f} | {s['avg_latency']*1000:.0f} ms |\n")

        total_row_faith = round(avg_faith, 4)
        f.write(f"| **TOPLAM** | **{total}** | **{total_passed}** | **%{accuracy:.2f}** | **{total_row_faith}** | **{avg_latency*1000:.0f} ms** |\n\n")
        f.write("---\n\n")
        f.write("## ❌ 3. Başarısız Sorular (İlk 20)\n\n")
        failed = [r for r in results if not r["passed"]]
        f.write(f"Toplam başarısız: **{len(failed)} soru**\n\n")
        for r in failed[:20]:
            f.write(f"### Q{r['qid']} — {r['category']}\n")
            f.write(f"**Soru:** {r['question'][:200]}\n\n")
            f.write(f"**Beklenen:** {r['expected'][:200]}\n\n")
            f.write(f"**Verilen:** {r['answer'][:300]}\n\n")
            f.write(f"**Not:** {r['notes']} | Gecikme: {r['latency_s']*1000:.0f}ms | Sadakat: {r['faithfulness']:.3f}\n\n")
            f.write("---\n\n")
        f.write(f"\n## 📁 4. Veri Dosyaları\n\n")
        f.write(f"- **Sorular:** `Microsoft_EcoRAG_Benchmark_Test_Set_2.xlsx`\n")
        f.write(f"- **Detaylı Loglar:** `benchmark_set2_results.json`\n")
        f.write(f"- **Bu Rapor:** `BENCHMARK_SET2_REPORT.md`\n")

    print(f"  Report  saved → {OUT_REPORT}")

    # ── Print summary ─────────────────────────────────────────────────────────
    print(f"\n{'='*72}")
    print(f"  SONUÇ: %{accuracy:.2f} genel doğruluk ({total_passed}/{total} başarılı)")
    print(f"  Ort. gecikme: {avg_latency*1000:.0f} ms | Süre: {total_time:.1f}s")
    print(f"  Ort. sadakat: {avg_faith:.4f}")
    print(f"{'='*72}")
    for cat in cat_order:
        if cat in summary["category_breakdown"]:
            s = summary["category_breakdown"][cat]
            print(f"  {cat[:42]:<42}  %{s['accuracy']:>6.2f}  ({s['passed']}/{s['total']})")
    print(f"{'='*72}\n")

    return summary

if __name__ == "__main__":
    run_benchmark()
